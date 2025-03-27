import os
import sqlalchemy
from sqlalchemy import create_engine, Column, Integer, String, Float, MetaData, Table
from sqlalchemy.orm import declarative_base, sessionmaker
from sqlalchemy.dialects.postgresql import JSONB
from dotenv import load_dotenv
import xlrd
from openpyxl import load_workbook
import re

# Carica le variabili d'ambiente
load_dotenv()

# Definizione del dizionario di mappatura
column_mapping = {
    'Marca': 'marca',
    'marca': 'marca',
    'brand': 'marca',
    'Descrizione Marca': 'marca',
    
    'Modello': 'modello',
    'modello': 'modello',
    'model': 'modello',
    'Descrizione Versione': 'modello',
    'Descrizione Modello': 'modello',
    'Descrizione veicolo': 'modello',
    
    'Model Year': 'anno',
    'Anno': 'anno',
    'anno': 'anno',
    'year': 'anno',
    
    'Prezzo Veicolo': 'prezzo',
    'Prezzo': 'prezzo',
    'prezzo': 'prezzo',
    'price': 'prezzo',
    'Prezzo Listino Privil.': 'prezzo',
    'Prezzo Veicolo': 'prezzo',
    
    'Colore': 'colore',
    'colore': 'colore',
    'color': 'colore',
    
    'Targa': 'targa',
    'targa': 'targa',
    'plate': 'targa',
    
    'Km (vei. AZ.)': 'chilometraggio',
    'Chilometraggio': 'chilometraggio',
    'chilometraggio': 'chilometraggio',
    'km': 'chilometraggio',
    'kilometers': 'chilometraggio',
    'Km (vei. AZ.)': 'chilometraggio'
}

# Creazione della connessione al database PostgreSQL
db_params = {
    'username': os.getenv('DB_USER', 'postgres'),
    'password': os.getenv('DB_PASSWORD', 'postgres'),
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': os.getenv('DB_PORT', '5432'),
    'database': os.getenv('DB_NAME', 'novicars')
}

db_url = f"postgresql://{db_params['username']}:{db_params['password']}@{db_params['host']}:{db_params['port']}/{db_params['database']}"
engine = create_engine(db_url)
Base = declarative_base()

# Definizione del modello per la tabella principale delle auto
class Auto(Base):
    __tablename__ = 'auto'
    
    id = Column(Integer, primary_key=True)
    marca = Column(String(100), index=True)
    modello = Column(String(100), index=True)
    anno = Column(Integer, index=True)
    prezzo = Column(Float)
    colore = Column(String(200), index=True)
    targa = Column(String(20), index=True)
    chilometraggio = Column(Integer)
    data_importazione = Column(sqlalchemy.DateTime, default=sqlalchemy.func.now())
    
    def __repr__(self):
        return f"<Auto(id={self.id}, marca='{self.marca}', modello='{self.modello}', targa='{self.targa}')>"

# Tabella per il tracciamento delle importazioni
class ImportLog(Base):
    __tablename__ = 'import_log'
    
    id = Column(Integer, primary_key=True)
    file_name = Column(String(255))
    timestamp = Column(sqlalchemy.DateTime, default=sqlalchemy.func.now())
    records_imported = Column(Integer)
    success = Column(sqlalchemy.Boolean)
    error_message = Column(String(500), nullable=True)
    
    def __repr__(self):
        return f"<ImportLog(file='{self.file_name}', records={self.records_imported}, success={self.success})>"

# Creazione delle tabelle nel database
Base.metadata.create_all(engine)

def clean_numeric_value(value):
    """Pulisce e converte un valore numerico"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    # Rimuove caratteri non numerici e converte in float
    cleaned = re.sub(r'[^\d.,]', '', str(value))
    cleaned = cleaned.replace(',', '.')
    try:
        return float(cleaned)
    except ValueError:
        return None

def clean_string_value(value):
    """Pulisce un valore stringa"""
    if value is None:
        return None
    return str(value).strip()

def process_row_xlsx(row, file_mapping):
    """Processa una riga di un file xlsx"""
    auto_data = {}
    
    # Processa tutte le colonne mappate
    for col_idx, db_col in file_mapping.items():
        cell_value = row[col_idx - 1].value
        
        if cell_value is not None:
            try:
                # Converti il valore nel tipo appropriato
                if db_col == 'anno':
                    auto_data[db_col] = int(clean_numeric_value(cell_value)) if clean_numeric_value(cell_value) is not None else None
                elif db_col == 'prezzo':
                    auto_data[db_col] = clean_numeric_value(cell_value)
                elif db_col == 'chilometraggio':
                    auto_data[db_col] = int(clean_numeric_value(cell_value)) if clean_numeric_value(cell_value) is not None else None
                else:
                    auto_data[db_col] = clean_string_value(cell_value)
            except (ValueError, TypeError) as e:
                print(f"Errore nella conversione del valore per la colonna {db_col}: {cell_value}")
                continue
    
    return auto_data

def process_row_xls(row, file_mapping):
    """Processa una riga di un file xls"""
    auto_data = {}
    
    # Processa tutte le colonne mappate
    for col_idx, db_col in file_mapping.items():
        cell_value = row[col_idx - 1]  # Il valore è già una stringa o un numero
        
        if cell_value is not None:
            try:
                # Converti il valore nel tipo appropriato
                if db_col == 'anno':
                    auto_data[db_col] = int(clean_numeric_value(cell_value)) if clean_numeric_value(cell_value) is not None else None
                elif db_col == 'prezzo':
                    auto_data[db_col] = clean_numeric_value(cell_value)
                elif db_col == 'chilometraggio':
                    auto_data[db_col] = int(clean_numeric_value(cell_value)) if clean_numeric_value(cell_value) is not None else None
                else:
                    auto_data[db_col] = clean_string_value(cell_value)
            except (ValueError, TypeError) as e:
                print(f"Errore nella conversione del valore per la colonna {db_col}: {cell_value}")
                continue
    
    return auto_data

def import_excel_to_database(file_path):
    print(f"Importazione del file: {file_path}")
    
    # Crea la sessione di database
    Session = sessionmaker(bind=engine)
    session = Session()
    
    # Log dell'importazione
    import_log = ImportLog(file_name=os.path.basename(file_path), records_imported=0, success=False)
    session.add(import_log)
    
    try:
        # Determina il tipo di file
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.xlsx':
            # Carica il workbook Excel
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Ottieni le intestazioni
            headers = [cell.value for cell in ws[1]]
            
            # Stampa le colonne del file Excel per debug
            print("\nColonne presenti nel file Excel:")
            for col in headers:
                print(f"- {col}")
        
        # Mappatura delle colonne
        file_mapping = {}
            for idx, col in enumerate(headers, 1):
                if col:
                    col_normalized = str(col).lower().replace(' ', '_').replace('.', '_').replace('(', '').replace(')', '')
                    for excel_col, db_col in column_mapping.items():
                        excel_col_normalized = excel_col.lower().replace(' ', '_').replace('.', '_').replace('(', '').replace(')', '')
                        if excel_col_normalized == col_normalized:
                            file_mapping[idx] = db_col
                            print(f"Colonna mappata: {col} -> {db_col}")
                            break
            
            # Stampa le colonne dopo la mappatura
            print("\nColonne dopo la mappatura:")
            for idx, col in enumerate(headers, 1):
                print(f"- {col}")
            
            records_imported = 0
            records_updated = 0
            
            # Itera sulle righe del foglio
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                try:
                    # Processa la riga
                    auto_data = process_row_xlsx(row, file_mapping)
                    
                    # Verifica se abbiamo i dati minimi necessari (marca e modello)
                    marca = auto_data.get('marca', '')
                    modello = auto_data.get('modello', '')
                    
                    if marca and modello:
                        targa = auto_data.get('targa', '')
                        
                        if targa:
                            # Cerca se esiste già un record con la stessa targa
                            existing_auto = session.query(Auto).filter(Auto.targa == targa).first()
                            
                            if existing_auto:
                                # Aggiorna il record esistente
                                for key, value in auto_data.items():
                                    setattr(existing_auto, key, value)
                                records_updated += 1
                            else:
                                # Crea un nuovo record
                                new_auto = Auto(**auto_data)
                                session.add(new_auto)
                                records_imported += 1
                        else:
                            # Crea un nuovo record senza targa
                            new_auto = Auto(**auto_data)
                            session.add(new_auto)
                            records_imported += 1
                except Exception as e:
                    print(f"Errore nell'elaborazione della riga {row_idx}: {str(e)}")
                    continue
            
            wb.close()
            
        elif file_ext == '.xls':
            # Carica il workbook Excel
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            
            # Ottieni le intestazioni
            headers = [ws.cell_value(0, i) for i in range(ws.ncols)]
            
            # Stampa le colonne del file Excel per debug
            print("\nColonne presenti nel file Excel:")
            for col in headers:
                print(f"- {col}")
            
            # Mappatura delle colonne
            file_mapping = {}
            for idx, col in enumerate(headers, 1):
                if col:
                    col_normalized = str(col).lower().replace(' ', '_').replace('.', '_').replace('(', '').replace(')', '')
                    for excel_col, db_col in column_mapping.items():
                        excel_col_normalized = excel_col.lower().replace(' ', '_').replace('.', '_').replace('(', '').replace(')', '')
                        if excel_col_normalized == col_normalized:
                            file_mapping[idx] = db_col
                            print(f"Colonna mappata: {col} -> {db_col}")
                            break
            
            # Stampa le colonne dopo la mappatura
            print("\nColonne dopo la mappatura:")
            for idx, col in enumerate(headers, 1):
                print(f"- {col}")
            
            records_imported = 0
            records_updated = 0
            
            # Itera sulle righe del foglio
            for row_idx in range(1, ws.nrows):
                try:
                    # Ottieni i valori della riga
                    row = [ws.cell_value(row_idx, i) for i in range(ws.ncols)]
                    
                    # Processa la riga
                    auto_data = process_row_xls(row, file_mapping)
                    
                    # Verifica se abbiamo i dati minimi necessari (marca e modello)
                    marca = auto_data.get('marca', '')
                    modello = auto_data.get('modello', '')
                    
                    if marca and modello:
                        targa = auto_data.get('targa', '')
                        
                        if targa:
                            # Cerca se esiste già un record con la stessa targa
                            existing_auto = session.query(Auto).filter(Auto.targa == targa).first()
                            
                            if existing_auto:
                                # Aggiorna il record esistente
                                for key, value in auto_data.items():
                                    setattr(existing_auto, key, value)
                                records_updated += 1
                            else:
            # Crea un nuovo record
            new_auto = Auto(**auto_data)
            session.add(new_auto)
            records_imported += 1
                        else:
                            # Crea un nuovo record senza targa
                            new_auto = Auto(**auto_data)
                            session.add(new_auto)
                            records_imported += 1
                except Exception as e:
                    print(f"Errore nell'elaborazione della riga {row_idx + 1}: {str(e)}")
                    continue
        
        # Commit finale
        session.commit()
        
        # Aggiorna il log di importazione
        import_log.records_imported = records_imported + records_updated
        import_log.success = True
        session.commit()
        
        print(f"\nImportazione completata:")
        print(f"- {records_imported} nuovi record importati")
        print(f"- {records_updated} record aggiornati")
        print(f"- Totale: {records_imported + records_updated} record elaborati")
        return records_imported + records_updated
    
    except Exception as e:
        session.rollback()
        error_msg = str(e)
        print(f"Errore durante l'importazione: {error_msg}")
        
        # Registra l'errore nel log
        import_log.error_message = error_msg[:500]  # Tronca se troppo lungo
        session.commit()
        
        return 0
    
    finally:
        session.close()

# Funzione principale per elaborare tutti i file in una directory
def process_all_excel_files(directory_path):
    total_imported = 0
    
    # Verifica che la directory esista
    if not os.path.exists(directory_path):
        print(f"La directory {directory_path} non esiste.")
        return
    
    # Elabora tutti i file Excel nella directory
    for filename in os.listdir(directory_path):
        if filename.endswith(('.xlsx', '.xls')):
            file_path = os.path.join(directory_path, filename)
            imported = import_excel_to_database(file_path)
            total_imported += imported
    
    print(f"Importazione completata. Totale record importati: {total_imported}")

# Esempio di utilizzo
if __name__ == "__main__":
    file_path = r"D:\Coding\novicars\Stock Audi Magnifica al 17.03.xls"
    import_excel_to_database(file_path)

